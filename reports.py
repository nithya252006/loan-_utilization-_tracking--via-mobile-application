import os
import tempfile
from datetime import datetime

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


REPORT_DIR = os.path.join(tempfile.gettempdir(), "loan_reports")
os.makedirs(REPORT_DIR, exist_ok=True)


def build_pdf_report(username, loans):
    filepath = os.path.join(REPORT_DIR, f"{username}_loan_report.pdf")
    doc = SimpleDocTemplate(filepath, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = [
        Paragraph(f"Loan Utilization Report — {username}", styles["Title"]),
        Spacer(1, 0.5 * cm),
    ]

    data = [["ID", "Amount", "Purpose", "Date", "Status", "Used", "Remaining", "Util %"]]
    for loan in loans:
        data.append([
            loan["id"],
            f"{float(loan['amount']):.2f}",
            loan["purpose"],
            str(loan["loan_date"]),
            loan["status"],
            f"{loan['used_amount']:.2f}",
            f"{loan['remaining_amount']:.2f}",
            f"{loan['utilization_pct']}%",
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#007BFF")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
    ]))
    elements.append(table)

    total_amount = sum(float(loan["amount"]) for loan in loans)
    elements.append(Spacer(1, 0.7 * cm))
    elements.append(Paragraph(f"Total Loan Amount: ₹{total_amount:.2f}", styles["Normal"]))
    elements.append(Paragraph(f"Username: {username}", styles["Normal"]))
    elements.append(Paragraph(
        f"Generated on: {datetime.now().strftime('%d-%m-%Y %H:%M')}", styles["Normal"]
    ))

    doc.build(elements)
    return filepath


def build_excel_report(username, loans):
    filepath = os.path.join(REPORT_DIR, f"{username}_loan_report.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Loan Report"

    headers = ["ID", "Amount", "Purpose", "Date", "Status", "Used", "Remaining", "Utilization %"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="007BFF", end_color="007BFF", fill_type="solid")

    for loan in loans:
        ws.append([
            loan["id"],
            float(loan["amount"]),
            loan["purpose"],
            str(loan["loan_date"]),
            loan["status"],
            loan["used_amount"],
            loan["remaining_amount"],
            loan["utilization_pct"],
        ])

    total_amount = sum(float(loan["amount"]) for loan in loans)
    ws.append([])
    ws.append(["Total Loan Amount:", total_amount])
    ws.append(["Username:", username])
    ws.append(["Generated on:", datetime.now().strftime("%d-%m-%Y %H:%M")])

    for col in ws.columns:
        max_len = max(len(str(c.value)) for c in col if c.value is not None) if col else 10
        ws.column_dimensions[col[0].column_letter].width = max(12, max_len + 2)

    wb.save(filepath)
    return filepath